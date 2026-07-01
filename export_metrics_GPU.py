import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time

PROMETHEUS = "http://localhost:9090"
HOURS = 4
STEP = "10s"

SECTIONS = {
    "CPU": ["node_cpu", "node_pressure_cpu", "node_schedstat", "node_softirq"],
    "Temperature": ["node_thermal", "node_hwmon", "node_cooling"],
    "Memory": ["node_mem", "node_vmstat", "node_pressure_mem"],
    "Disk": ["node_disk", "node_filesystem", "node_pressure_io"],
    "Network": ["node_network"],
    "System": ["node_load", "node_boot", "node_time", "node_uname", "node_os",
               "node_procs", "node_intr", "node_context", "node_entropy", "node_forks"],
    "Power": ["node_power", "node_rapl"],
    "GPU": ["nvidia_smi", "nvidia_gpu"],
}

MAX_ROWS = 1048575  # Excel лимит минус заголовок

def get_all_metrics():
    r = requests.get(f"{PROMETHEUS}/api/v1/label/__name__/values")
    r.raise_for_status()
    all_metrics = r.json()["data"]
    return [m for m in all_metrics if m.startswith("node_") or m.startswith("nvidia_")]

def assign_section(metric_name):
    for section, prefixes in SECTIONS.items():
        for prefix in prefixes:
            if metric_name.startswith(prefix):
                return section
    return "Other"

def fetch_metric(metric_name, start, end):
    params = {
        "query": metric_name,
        "start": start,
        "end": end,
        "step": STEP,
    }
    r = requests.get(f"{PROMETHEUS}/api/v1/query_range", params=params)
    if r.status_code != 200:
        return None
    data = r.json().get("data", {}).get("result", [])
    if not data:
        return None

    rows = []
    for series in data:
        labels = series["metric"]
        label_str = ", ".join(f"{k}={v}" for k, v in labels.items() if k != "__name__")
        for ts, val in series["values"]:
            try:
                rows.append({
                    "time": datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S"),
                    "metric": metric_name,
                    "labels": label_str,
                    "value": float(val),
                })
            except (ValueError, TypeError):
                pass
    return rows if rows else None

def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color="2E4057")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

def write_sheet(wb, section_name, all_rows):
    chunks = [all_rows[i:i+MAX_ROWS] for i in range(0, len(all_rows), MAX_ROWS)]
    for idx, chunk in enumerate(chunks):
        title = section_name[:31] if idx == 0 else f"{section_name[:28]}_{idx+1}"
        ws = wb.create_sheet(title=title)
        headers = ["Time", "Metric", "Labels", "Value"]
        ws.append(headers)
        style_header(ws, len(headers))
        for row in chunk:
            ws.append([row["time"], row["metric"], row["labels"], row["value"]])
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 15
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal="left")

def main():
    end = time.time()
    start = end - HOURS * 3600

    print("Получаю список метрик...")
    all_metrics = get_all_metrics()
    print(f"Найдено метрик: {len(all_metrics)}")

    by_section = {s: [] for s in list(SECTIONS.keys()) + ["Other"]}

    for i, metric in enumerate(all_metrics):
        section = assign_section(metric)
        print(f"[{i+1}/{len(all_metrics)}] {metric} → {section}")
        rows = fetch_metric(metric, start, end)
        if rows:
            by_section[section].extend(rows)

    wb = Workbook()
    wb.remove(wb.active)

    # Сводный лист
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["Section", "Rows"])
    style_header(ws_summary, 2)
    for section, rows in by_section.items():
        ws_summary.append([section, len(rows)])
    ws_summary.append(["TOTAL", sum(len(r) for r in by_section.values())])
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15

    for section, rows in by_section.items():
        if rows:
            print(f"Записываю лист: {section} ({len(rows)} строк)")
            write_sheet(wb, section, rows)

    filename = f"prometheus_metrics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"\nГотово! Файл сохранён: {filename}")

if __name__ == "__main__":
    main()
